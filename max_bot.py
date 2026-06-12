import asyncio
import logging
import os
import aiohttp
from datetime import datetime
from openpyxl import load_workbook
import re

# Настройки
# Читаем токены
MAX_TOKEN = os.getenv("MAX_TOKEN", "")
TG_BOT_TOKEN = os.getenv("BOT_TOKEN", "")
DRIVERS_CHAT_ID = int(os.getenv("DRIVERS_CHAT_ID", "0"))

MAX_API = "https://platform-api.max.ru"
TG_API = f"https://api.telegram.org/bot{TG_BOT_TOKEN}"

# Логируем что загрузилось
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logger.info(f"MAX_TOKEN: '{MAX_TOKEN[:15]}...' (длина: {len(MAX_TOKEN)})")
logger.info(f"BOT_TOKEN: '{TG_BOT_TOKEN[:10]}...'")
logger.info(f"DRIVERS_CHAT_ID: {DRIVERS_CHAT_ID}")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Хранилища
max_orders = {}
max_order_counter = 0
max_states = {}   # user_id -> {"step": ..., "from": ..., "order_id": ...}
max_marker = 0

# Загружаем прайс
PRICE_TABLE = {}
def load_prices():
    global PRICE_TABLE
    try:
        wb = load_workbook("price.xlsx", read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue
            if row[0] and row[2] and row[4]:
                f = str(row[0]).strip().lower()
                t = str(row[2]).strip().lower()
                PRICE_TABLE[(f, t)] = (int(row[4]), int(row[5]) if row[5] else int(row[4]) + 50)
        logger.info(f"Загружено маршрутов: {len(PRICE_TABLE)}")
    except Exception as e:
        logger.error(f"Ошибка загрузки прайса: {e}")

load_prices()


def normalize(addr):
    addr = addr.lower().strip()
    addr = re.sub(r'\s+\d+[а-яa-z]?$', '', addr)
    addr = re.sub(r',.*', '', addr)
    addr = re.sub(r'^(ул\.|улица|пер\.|переулок)\s+', '', addr)
    return addr.strip()


def get_price(from_addr, to_addr):
    fn = normalize(from_addr)
    tn = normalize(to_addr)
    now = datetime.now()
    is_night = now.hour < 6 or (now.hour == 6 and now.minute < 30)

    if (fn, tn) in PRICE_TABLE:
        d, n = PRICE_TABLE[(fn, tn)]
        price = n if is_night else d
        night = " (ночной тариф 🌙)" if is_night else ""
        return f"💰 Предварительная стоимость: {price} руб.{night}\nТочную стоимость укажет водитель."

    for (f, t), (d, n) in PRICE_TABLE.items():
        if (fn in f or f in fn) and (tn in t or t in tn):
            price = n if is_night else d
            night = " (ночной тариф 🌙)" if is_night else ""
            return f"💰 Предварительная стоимость: {price} руб.{night}\nТочную стоимость укажет водитель."

    return "💰 Предварительная стоимость: от 150 руб.\nТочную стоимость укажет водитель."


# ─── ОТПРАВКА В МАКС ─────────────────────────

async def send_max(session, chat_id, text):
    url = f"{MAX_API}/messages"
    headers = {"Authorization": MAX_TOKEN}
    payload = {
        "recipient": {"chat_id": int(chat_id)},
        "body": {"text": text}
    }
    try:
        async with session.post(url, headers=headers, json=payload) as r:
            result = await r.json()
            logger.info(f"MAX send: {result}")
            return result
    except Exception as e:
        logger.error(f"Ошибка отправки в Макс: {e}")


# ─── ОТПРАВКА В TELEGRAM ──────────────────────

async def send_tg(session, chat_id, text, keyboard=None):
    url = f"{TG_API}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if keyboard:
        payload["reply_markup"] = keyboard
    try:
        async with session.post(url, json=payload) as r:
            return await r.json()
    except Exception as e:
        logger.error(f"Ошибка отправки в Telegram: {e}")


# ─── УВЕДОМЛЕНИЕ 5 МИНУТ ─────────────────────

async def notify_no_driver(session, order_id, client_chat_id):
    await asyncio.sleep(300)
    if order_id not in max_orders:
        return
    if max_orders[order_id]["status"] != "new":
        return
    await send_max(session, client_chat_id,
        "⚠️ Пока никто не взял ваш заказ.\n\n"
        "Напишите /start чтобы попробовать снова.\n"
        "Приносим извинения! 🙏"
    )


# ─── ОБРАБОТКА СООБЩЕНИЙ ─────────────────────

async def handle_update(session, update):
    global max_order_counter

    update_type = update.get("update_type")
    logger.info(f"Получено обновление: {update_type}")

    # Получаем сообщение
    if update_type == "message_created":
        message = update.get("message", {})
    elif update_type == "bot_started":
        message = update.get("message", {})
    else:
        return

    logger.info(f"ПОЛНЫЙ UPDATE: {update}")

    sender = message.get("sender", {})
    user_id = sender.get("user_id")

    # В MAX API личные сообщения нужно отправлять с chat_id, не user_id
    recipient = message.get("recipient", {})
    chat_id = recipient.get("chat_id") or recipient.get("user_id") or user_id

    body = message.get("body", {})
    text = body.get("text", "").strip() if body else ""

    if not user_id:
        return

    logger.info(f"Сообщение от user_id={user_id}, chat_id={chat_id}: {text}")

    state = max_states.get(user_id, {})

    # /start или bot_started
    if text.lower() in ["/start", "start", "привет"] or update_type == "bot_started":
        max_states[user_id] = {"step": "waiting_from", "chat_id": chat_id}
        await send_max(session, chat_id,
            "🚕 Добро пожаловать в службу такси!\n\n"
            "Введите адрес ОТКУДА вас забрать:"
        )
        return

    # Шаг 1 — откуда
    if state.get("step") == "waiting_from":
        max_states[user_id] = {"step": "waiting_to", "from": text, "chat_id": chat_id}
        await send_max(session, chat_id,
            f"📍 Откуда: {text}\n\n"
            f"Теперь введите адрес КУДА вас везти:"
        )
        return

    # Шаг 2 — куда
    if state.get("step") == "waiting_to":
        from_addr = state["from"]
        to_addr = text

        max_order_counter += 1
        order_id = f"M{max_order_counter}"

        max_orders[order_id] = {
            "client_id": user_id,
            "client_chat_id": chat_id,
            "from_addr": from_addr,
            "to_addr": to_addr,
            "driver_tg_id": None,
            "status": "new",
            "date": datetime.now().strftime("%d.%m.%Y %H:%M")
        }
        max_states[user_id] = {"step": "active", "order_id": order_id, "chat_id": chat_id}

        price_text = get_price(from_addr, to_addr)

        await send_max(session, chat_id,
            f"✅ Ваш заказ принят!\n\n"
            f"📍 Откуда: {from_addr}\n"
            f"🏁 Куда: {to_addr}\n\n"
            f"{price_text}\n\n"
            f"⏳ Ищем водителя, ожидайте..."
        )

        # Отправляем в чат водителей Telegram
        keyboard = {
            "inline_keyboard": [[
                {"text": "✅ Взять заказ (MAX)", "callback_data": f"takemax_{order_id}"}
            ]]
        }
        await send_tg(session, DRIVERS_CHAT_ID,
            f"🚖 *НОВЫЙ ЗАКАЗ {order_id}* (из Макса 📱)\n\n"
            f"📍 Откуда: {from_addr}\n"
            f"🏁 Куда: {to_addr}\n",
            keyboard=keyboard
        )

        asyncio.create_task(notify_no_driver(session, order_id, chat_id))
        return

    # Переписка клиента с водителем
    if state.get("step") == "active":
        order_id = state.get("order_id")
        active_chat_id = state.get("chat_id", chat_id)
        if order_id and order_id in max_orders:
            order = max_orders[order_id]
            if order.get("driver_tg_id"):
                await send_tg(session, order["driver_tg_id"],
                    f"💬 *Сообщение от клиента (MAX):*\n{text}"
                )
            else:
                await send_max(session, active_chat_id, "⏳ Ожидайте, водитель ещё не найден...")
        return

    # Если ничего не подошло
    await send_max(session, chat_id,
        "Напишите /start чтобы заказать такси 🚕"
    )


# ─── ОСНОВНОЙ ЦИКЛ ───────────────────────────

async def polling():
    global max_marker

    token = os.getenv("MAX_TOKEN", "NOT_FOUND")
    logger.info(f"Токен из env: длина={len(token)}, начало={token[:10]}")

    # Пробуем разные форматы авторизации
    headers_variants = [
        {"Authorization": token},
        {"Authorization": f"Bearer {token}"},
        {"access-token": token},
    ]

    async with aiohttp.ClientSession() as session:
        logger.info("MAX бот запущен! Начинаем polling...")

        # Проверяем подключение всеми способами
        for headers in headers_variants:
            try:
                async with session.get(f"{MAX_API}/me", headers=headers) as r:
                    me = await r.json()
                    logger.info(f"Попытка с {list(headers.keys())[0]}: {me}")
                    if me.get("user_id") or me.get("name"):
                        logger.info(f"✅ Успешно! Используем: {list(headers.keys())[0]}")
                        break
            except Exception as e:
                logger.error(f"Ошибка: {e}")

        headers = {"Authorization": token}

        while True:
            try:
                params = {"timeout": 20, "limit": 100}
                if max_marker:
                    params["marker"] = max_marker

                async with session.get(
                    f"{MAX_API}/updates",
                    headers=headers,
                    params=params,
                    timeout=aiohttp.ClientTimeout(total=30)
                ) as r:
                    data = await r.json()

                if "marker" in data:
                    max_marker = data["marker"]

                updates = data.get("updates", [])
                for update in updates:
                    try:
                        await handle_update(session, update)
                    except Exception as e:
                        logger.error(f"Ошибка обработки: {e}")

            except asyncio.TimeoutError:
                pass
            except Exception as e:
                logger.error(f"Ошибка polling: {e}")
                await asyncio.sleep(5)


if __name__ == "__main__":
    asyncio.run(polling())
